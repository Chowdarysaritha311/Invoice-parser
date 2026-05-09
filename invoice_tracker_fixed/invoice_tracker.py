import os, requests, base64, json, re, datetime, openpyxl, smtplib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.text import MIMEText
from pdf2image import convert_from_path
from PIL import Image

# ── Settings ──────────────────────────────────────
MODEL        = 'qwen/qwen2.5-vl-7b'
LM_URL       = 'http://127.0.0.1:1234/v1/chat/completions'
TDS_API_URL  = 'http://localhost:5000/tds-rules'
EXCEL_FILE   = os.path.join(os.path.dirname(__file__), 'invoice_tracker.xlsx')
IMAGES_DIR   = os.path.join(os.path.dirname(__file__), 'invoice_pages')
POPPLER_PATH = r'C:\Users\chowd\Downloads\Release-25.12.0-0\poppler-25.12.0\Library\bin'

EMAIL_ENABLED  = False
EMAIL_SENDER   = 'your@gmail.com'
EMAIL_PASSWORD = 'your_app_password'
EMAIL_TO       = 'finance@company.com'
# ──────────────────────────────────────────────────


# ── AI helpers ────────────────────────────────────

def ai(messages, image=None, tokens=600):
    if image:
        messages[-1]['content'] = [
            {'type': 'image_url', 'image_url': {'url': f'data:image/png;base64,{image}'}},
            {'type': 'text', 'text': messages[-1]['content']}
        ]
    r = requests.post(LM_URL, json={'model': MODEL, 'messages': messages,
                      'temperature': 0.0, 'max_tokens': tokens, 'stream': False}, timeout=600)
    return r.json()['choices'][0]['message']['content'].strip()


def parse(text):
    text = re.sub(r'```json|```', '', text).strip()
    try: return json.loads(text)
    except:
        m = re.search(r'\{.*\}', text, re.DOTALL)
        return json.loads(m.group()) if m else {}


def tds_rules():
    try: return requests.get(TDS_API_URL, timeout=10).json()['rules']
    except:
        f = os.path.join(os.path.dirname(__file__), 'tds_rules.txt')
        return open(f, encoding='utf-8').read() if os.path.exists(f) else ''


# ── Utility helpers ───────────────────────────────

def fv(data, key):
    v = data.get(key, '')
    return '' if v in (None, 'NOT AVAILABLE', 'null', 'N/A') else str(v).strip()


def ask(label, default=''):
    default = str(default) if default else ''
    v = input(f'  {label:<20} [{default or "empty"}]: ').strip()
    return v if v else default


def parse_date(d):
    for f in ['%d %b %Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d %B %Y']:
        try: return datetime.datetime.strptime(d.strip(), f)
        except: pass
    return None


def month_year(d):
    dt = parse_date(d)
    return (dt.strftime('%B'), str(dt.year)) if dt else ('', '')


# ── PDF to image ──────────────────────────────────

def pdf_to_images(pdf_path):
    os.makedirs(IMAGES_DIR, exist_ok=True)
    pages = convert_from_path(pdf_path, dpi=100, poppler_path=POPPLER_PATH)
    images = []
    for i, page in enumerate(pages):
        w, h = page.size
        if w > 480: page = page.resize((480, int(h*480/w)), Image.LANCZOS)
        path = os.path.join(IMAGES_DIR, f'page_{i+1}.png')
        page.save(path, 'PNG')
        images.append(base64.b64encode(open(path,'rb').read()).decode('utf-8'))
    return images


# ── Invoice extraction ────────────────────────────

EXTRACT_PROMPT = (
    'Read this invoice and return ONLY this JSON:\n'
    '{"document_type":"","vendor_name":"","vendor_gstin":"","invoice_number":"",'
    '"invoice_date":"","due_date":"","buyer_name":"","buyer_gstin":"","hsn_code":"",'
    '"service_description":"","material_indicator":"","base_amount":"","cgst":"",'
    '"sgst":"","igst":"","total_with_gst":""}\n'
    'RULES:\n'
    '- vendor=company at TOP of invoice. buyer=Bill To section.\n'
    '- Dates: dd mmm yyyy format only (e.g. 30 Jan 2026)\n'
    '- Amounts: plain numbers only. NO Rs INR commas or currency symbols. e.g. 1500.00\n'
    '- cgst/sgst/igst: extract the RUPEE AMOUNT only not the percentage. e.g. 135.00 not 9%\n'
    '- If cgst/sgst/igst only shows a percentage and no rupee amount leave it empty\n'
    '- base_amount: single final number only. No formulas no plus signs. e.g. 1000.00\n'
    '- igst: empty string if not present\n'
    '- invoice_number: exact number from invoice. Do not add prefixes.'
)


def clean_amount(val):
    if not val:
        return ''
    s = str(val).strip()
    # If only a percentage like 9% return empty
    if re.match(r'^\d+(\.\d+)?%$', s.strip()):
        return ''
    # Remove currency symbols
    s = s.replace('\u20b9', '').replace('Rs', '').replace('INR', '').replace('$', '').strip()
    # If formula with = extract last number
    if '=' in s:
        s = s.split('=')[-1].strip()
    # If has + signs sum them
    if '+' in s and '=' not in s:
        try:
            nums = re.findall(r'[\d.]+', s)
            return str(round(sum(float(n) for n in nums), 2))
        except:
            pass
    # Remove commas
    s = s.replace(',', '').strip()
    # Extract first valid number
    m = re.search(r'[\d.]+', s)
    return m.group() if m else ''


def clean_extracted(data):
    for f in ['base_amount', 'cgst', 'sgst', 'igst', 'total_with_gst']:
        if f in data:
            data[f] = clean_amount(data[f])
    return data


def extract_invoice(pdf_path):
    print(f'Reading: {pdf_path}')
    images  = pdf_to_images(pdf_path)
    result  = {}
    for i, img in enumerate(images):
        print(f'  Page {i+1} of {len(images)}...')
        data = parse(ai([{'role': 'user', 'content': EXTRACT_PROMPT}], image=img))
        for k, v in data.items():
            if v and v not in ('', 'NOT AVAILABLE', 'null', 'N/A'):
                result.setdefault(k, v)
    return clean_extracted(result)


# ── TDS decision ──────────────────────────────────

def decide_tds(extracted):
    """Fully automatic TDS decision - no human input needed.
    AI checks HSN code, material indicator, and service description
    against TDS rules and decides automatically."""
    result = parse(ai([
        {'role': 'system', 'content': tds_rules()},
        {'role': 'user', 'content':
            f'Vendor: {fv(extracted,"vendor_name")}\n'
            f'Service: {fv(extracted,"service_description")}\n'
            f'HSN Code: {fv(extracted,"hsn_code")}\n'
            f'Materials on invoice: {fv(extracted,"material_indicator")}\n'
            f'Amount: Rs {fv(extracted,"base_amount")}\n\n'
            'Based on the TDS rules, decide:\n'
            '1. Does the vendor supply all materials needed for this work? (check HSN and material indicator)\n'
            '2. If yes - No TDS applies (product purchase)\n'
            '3. If no or unclear - apply correct TDS section\n'
            'Reply ONLY as JSON: {"tds_value":"","reason":"","confidence":"high/medium/low"}'}
    ], tokens=150))
    tds   = result.get('tds_value', 'Unable to determine')
    why   = result.get('reason', '')
    conf  = result.get('confidence', 'low')
    return tds, why, conf


def decide_tds_manual(extracted):
    """Manual TDS decision with human input - used in single invoice flow."""
    print(f'\n  Service  : {fv(extracted,"service_description")}')
    print(f'  HSN Code : {fv(extracted,"hsn_code")}')
    print(f'  Materials: {fv(extracted,"material_indicator")}')
    print('  Y=Vendor supplies all materials (No TDS)  N=We supply (TDS applies)  S=Skip')
    ans = input('  Answer (Y/N/S): ').strip().upper()
    if ans == 'Y':
        return 'No TDS', 'Vendor supplies all materials', 'high'
    elif ans == 'N':
        result = parse(ai([
            {'role': 'system', 'content': tds_rules()},
            {'role': 'user', 'content':
                f'Vendor:{fv(extracted,"vendor_name")} Service:{fv(extracted,"service_description")} '
                f'HSN:{fv(extracted,"hsn_code")} Amount:Rs {fv(extracted,"base_amount")}\n'
                'Reply ONLY as JSON: {"tds_value":"","reason":"","confidence":"high/medium/low"}'}
        ], tokens=150))
        return result.get('tds_value',''), result.get('reason',''), result.get('confidence','low')
    return '', 'Skipped', 'low'


# ── Smart checks ──────────────────────────────────

def check_duplicate(vendor, invoice_no):
    if not os.path.exists(EXCEL_FILE): return False
    ws = openpyxl.load_workbook(EXCEL_FILE).active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if (str(row[12] or '').lower() == vendor.lower() and
                str(row[14] or '').lower() == invoice_no.lower()):
            return True
    return False


def check_due_date(due_str):
    dt = parse_date(due_str)
    if not dt: return
    diff = (datetime.datetime.today() - dt).days
    if diff > 0:   print(f'  WARNING: Payment was due {diff} day(s) ago!')
    elif diff == 0: print(f'  WARNING: Payment due TODAY!')
    elif diff >= -3: print(f'  REMINDER: Payment due in {abs(diff)} day(s)')


def confidence_score(data):
    keys   = ['vendor_name','invoice_number','invoice_date','base_amount','total_with_gst','hsn_code']
    filled = sum(1 for k in keys if fv(data, k))
    score  = filled / len(keys)
    if score >= 0.8:   return 'High',   []
    elif score >= 0.5: return 'Medium', [k for k in keys if not fv(data, k)]
    else:              return 'Low',    [k for k in keys if not fv(data, k)]


# ── Excel helpers ─────────────────────────────────

COLS = [
    'SI No', 'Week', 'Month Name', 'Year', 'ID', 'Invoice / PI',
    'Invoice Date', 'Due Date', 'Project Name', 'Owner', 'Category',
    'Vendor ID', 'Vendor Name', 'TDS', 'Invoice Number', 'INV_SaveCopy',
    'Inv_Amt', 'Amount Paid / To be Paid', 'Partial Paid Amount',
    'Payment Status', 'Paid From', 'Payment Date',
    'GST Inputs Received Or Not', 'Material Received Date',
    'Material Received Status', 'Base Amount', 'SGST', 'CGST', 'IGST',
    'TDS Payment Date', 'TDS Done', 'Comments'
]
WIDTHS = [
    8, 10, 14, 8, 20, 15,
    14, 14, 20, 15, 15,
    12, 25, 15, 18, 55,
    16, 20, 18,
    16, 15, 14,
    22, 20,
    20, 14, 12, 12, 12,
    16, 12, 20
]


def get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        return wb, wb.active

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Invoice Tracker'

    # Plain bold header - no colors, just like normal Excel
    for i, c in enumerate(COLS, 1):
        cell = ws.cell(1, i, c)
        cell.font      = Font(name='Calibri', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'
    ws.row_dimensions[1].height = 30

    wb.save(EXCEL_FILE)
    return wb, ws


def build_row(data, si):
    mn, yr = month_year(data.get('invoice_date',''))
    uid    = data.get('unique_id','')
    dt     = data.get('invoice_date','')
    v      = data.get('vendor_name','')
    inv    = data.get('invoice_number','')
    tot    = data.get('total_amount','')
    copy   = f"{uid} {dt} {v} {inv} {tot}"
    return [
        si,
        data.get('week',''),
        mn, yr, uid,
        data.get('doc_type',''),
        dt,
        data.get('due_date',''),
        data.get('project_name',''),
        data.get('owner',''),
        data.get('category',''),
        data.get('vendor_id',''),
        v,
        data.get('tds',''),
        inv,
        copy,
        tot,
        data.get('amount_paid',''),
        data.get('partial_paid',''),
        data.get('payment_status',''),
        data.get('paid_from',''),
        data.get('payment_date',''),
        data.get('gst_inputs',''),
        data.get('material_received_date',''),
        data.get('material_received_status',''),
        data.get('base_amount',''),
        data.get('sgst',''),
        data.get('cgst',''),
        data.get('igst',''),
        data.get('tds_payment_date',''),
        data.get('tds_done',''),
        data.get('comments',''),
    ], copy


def write_row(ws, row_no, values):
    for i, val in enumerate(values, 1):
        cell = ws.cell(row_no, i, val)
        cell.font      = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row_no].height = 18


def save_invoice(data):
    wb, ws  = get_or_create_workbook()
    row_no  = ws.max_row + 1
    si      = ws.max_row - 1  # subtract title row and header row
    values, copy = build_row(data, si)
    write_row(ws, row_no, values)
    wb.save(EXCEL_FILE)
    print(f'Saved! Row {si} | File: {EXCEL_FILE}')
    print(f'Invoice Copy: {copy}')


# ── Email ─────────────────────────────────────────

def send_email(data):
    if not EMAIL_ENABLED: return
    try:
        body = (f"Invoice Saved\n\nVendor: {data.get('vendor_name')}\n"
                f"Invoice No: {data.get('invoice_number')}\nDate: {data.get('invoice_date')}\n"
                f"Amount: Rs {data.get('total_amount')}\nTDS: {data.get('tds')}\n"
                f"Due Date: {data.get('due_date')}")
        msg            = MIMEText(body)
        msg['Subject'] = f"New Invoice - {data.get('vendor_name')} - {data.get('invoice_number')}"
        msg['From']    = EMAIL_SENDER
        msg['To']      = EMAIL_TO
        s = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        s.login(EMAIL_SENDER, EMAIL_PASSWORD)
        s.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())
        s.quit()
        print('Email sent.')
    except Exception as e:
        print(f'Email failed: {e}')


# ── Review form ───────────────────────────────────

def review_form(extracted, tds_val, unique_id):
    print('\nREVIEW — Press ENTER to keep, type to correct')
    print('-' * 45)
    return {
        'unique_id':      ask('Unique ID',      unique_id),
        'doc_type':       ask('Document Type',  fv(extracted,'document_type')),
        'vendor_name':    ask('Vendor Name',    fv(extracted,'vendor_name')),
        'vendor_gstin':   ask('Vendor GSTIN',   fv(extracted,'vendor_gstin')),
        'invoice_number': ask('Invoice Number', fv(extracted,'invoice_number')),
        'invoice_date':   ask('Invoice Date',   fv(extracted,'invoice_date')),
        'due_date':       ask('Due Date',       fv(extracted,'due_date')),
        'buyer_name':     ask('Buyer Name',     fv(extracted,'buyer_name')),
        'hsn_code':       ask('HSN Code',       fv(extracted,'hsn_code')),
        'service_desc':   ask('Service',        fv(extracted,'service_description')),
        'base_amount':    ask('Base Amount',    fv(extracted,'base_amount')),
        'cgst':           ask('CGST',           fv(extracted,'cgst')),
        'sgst':           ask('SGST',           fv(extracted,'sgst')),
        'igst':           ask('IGST',           fv(extracted,'igst')),
        'total_amount':   ask('Total Amount',   fv(extracted,'total_with_gst')),
        'tds':            ask('TDS',            tds_val),
        'week':           ask('Week',           ''),
        'payment_status': ask('Payment Status', ''),
        'paid_from':      ask('Paid From',      ''),
        'payment_date':   ask('Payment Date',   ''),
        'tds_done':       ask('TDS Done',       ''),
    }


# ══════════════════════════════════════════════════
# FLOW 1 — ADD SINGLE INVOICE
# ══════════════════════════════════════════════════

def flow_add_invoice():
    pdf_name  = input('\nPDF file name: ').strip()
    unique_id = input('Unique ID    : ').strip() or 'AUTO'
    pdf_path  = os.path.join(os.path.dirname(__file__), pdf_name)
    if not os.path.exists(pdf_path):
        print('File not found.'); return

    extracted = extract_invoice(pdf_path)
    if not extracted:
        print('Could not read invoice.'); return

    conf, missing = confidence_score(extracted)
    print(f'\nConfidence: {conf}')
    if missing: print(f'Check: {", ".join(missing)}')

    print('\nExtracted:')
    for k, v in extracted.items(): print(f'  {k}: {v}')

    if check_duplicate(fv(extracted,'vendor_name'), fv(extracted,'invoice_number')):
        print(f'\nWARNING: This invoice already exists!')
        if input('Continue? (Y/N): ').strip().upper() != 'Y': return

    if fv(extracted,'due_date'): check_due_date(fv(extracted,'due_date'))

    print('\nMATERIAL CHECK')
    tds_val, tds_why, tds_conf = decide_tds_manual(extracted)
    print(f'TDS: {tds_val} | {tds_why}')

    final = review_form(extracted, tds_val, unique_id)

    print('\nSUMMARY')
    for k, v in final.items(): print(f'  {k:<20}: {v or "(empty)"}')

    if input('\nSave? (Y/N): ').strip().upper() == 'Y':
        save_invoice(final)
        send_email(final)
        print('Done!')


# ══════════════════════════════════════════════════
# FLOW 2 — BATCH PROCESS MULTIPLE INVOICES
# Put all PDFs in a folder, script processes all one by one
# ══════════════════════════════════════════════════

def flow_batch():
    folder = input('\nFolder path containing PDFs: ').strip().strip('"')
    if not os.path.isdir(folder):
        print('Folder not found.'); return

    pdfs = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
    if not pdfs:
        print('No PDF files found in folder.'); return

    print(f'\nFound {len(pdfs)} PDF(s):')
    for i, p in enumerate(pdfs, 1): print(f'  {i}. {p}')

    if input('\nProcess all automatically? (Y/N): ').strip().upper() != 'Y': return

    # Results tracking
    saved    = []   # successfully saved
    skipped  = []   # low confidence / unreadable
    duplicates = [] # already exists in Excel

    print('\n' + '='*50)
    print('STARTING AUTOMATIC BATCH PROCESSING')
    print('No human input needed — system decides automatically')
    print('='*50)

    for pdf_name in pdfs:
        print(f'\n>>> {pdf_name}')
        pdf_path = os.path.join(folder, pdf_name)

        try:
            # Step 1: Extract invoice data
            extracted = extract_invoice(pdf_path)
            if not extracted:
                print(f'  SKIPPED — Could not read invoice (unreadable PDF)')
                skipped.append((pdf_name, 'Could not read invoice'))
                continue

            # Step 2: Check extraction quality
            conf, missing = confidence_score(extracted)
            print(f'  Confidence: {conf}')

            # Skip if confidence is Low — not enough data extracted
            if conf == 'Low':
                print(f'  SKIPPED — Low confidence. Missing: {", ".join(missing)}')
                skipped.append((pdf_name, f'Low confidence - missing: {", ".join(missing)}'))
                continue

            # Step 3: Check for duplicate
            vendor = fv(extracted, 'vendor_name')
            inv_no = fv(extracted, 'invoice_number')
            if check_duplicate(vendor, inv_no):
                print(f'  SKIPPED — Duplicate invoice already in Excel')
                duplicates.append((pdf_name, f'{vendor} - {inv_no}'))
                continue

            # Step 4: Auto TDS decision — no human needed
            print(f'  Deciding TDS automatically...')
            tds_val, tds_why, tds_conf = decide_tds(extracted)
            print(f'  TDS: {tds_val} ({tds_conf} confidence) | {tds_why}')

            # If TDS confidence is low, mark as needs review but still save
            tds_note = ''
            if tds_conf == 'low':
                tds_note = f'[NEEDS REVIEW] {tds_val}'
                tds_val  = tds_note

            # Step 5: Auto generate unique ID from invoice details
            uid = f"{fv(extracted,'invoice_number')}_{fv(extracted,'invoice_date').replace(' ','')}"
            if not uid.strip('_'): uid = 'AUTO'

            # Step 6: Build final data automatically
            final = {
                'unique_id':               uid,
                'doc_type':                fv(extracted, 'document_type'),
                'vendor_name':             vendor,
                'vendor_gstin':            fv(extracted, 'vendor_gstin'),
                'invoice_number':          inv_no,
                'invoice_date':            fv(extracted, 'invoice_date'),
                'due_date':                fv(extracted, 'due_date'),
                'buyer_name':              fv(extracted, 'buyer_name'),
                'buyer_gstin':             fv(extracted, 'buyer_gstin'),
                'hsn_code':                fv(extracted, 'hsn_code'),
                'service_desc':            fv(extracted, 'service_description'),
                'base_amount':             clean_amount(fv(extracted, 'base_amount')),
                'cgst':                    clean_amount(fv(extracted, 'cgst')),
                'sgst':                    clean_amount(fv(extracted, 'sgst')),
                'igst':                    clean_amount(fv(extracted, 'igst')),
                'total_amount':            clean_amount(fv(extracted, 'total_with_gst')),
                'tds':                     tds_val,
                'week':                    '',
                'project_name':            '',
                'owner':                   '',
                'category':                '',
                'vendor_id':               '',
                'amount_paid':             '',
                'partial_paid':            '',
                'payment_status':          '',
                'paid_from':               '',
                'payment_date':            '',
                'gst_inputs':              '',
                'material_received_date':  '',
                'material_received_status':'',
                'tds_payment_date':        '',
                'tds_done':                '',
                'comments':                f'Auto-processed. Confidence:{conf}. {tds_why}',
            }

            # Step 7: Check due date silently
            if fv(extracted, 'due_date'):
                dt = parse_date(fv(extracted, 'due_date'))
                if dt:
                    diff = (datetime.datetime.today() - dt).days
                    if diff > 0:
                        final['comments'] += f' | OVERDUE by {diff} days'

            # Step 8: Save to Excel automatically
            save_invoice(final)
            send_email(final)
            saved.append((pdf_name, vendor, inv_no, tds_val))
            print(f'  SAVED')

        except Exception as e:
            print(f'  ERROR: {e}')
            skipped.append((pdf_name, f'Error: {e}'))

    # Final report
    print('\n' + '='*50)
    print('BATCH COMPLETE')
    print('='*50)
    print(f'\nSaved     : {len(saved)}')
    for s in saved:
        print(f'  + {s[0]} | {s[1]} | {s[2]} | TDS: {s[3]}')

    print(f'\nSkipped   : {len(skipped)}')
    for s in skipped:
        print(f'  - {s[0]} | Reason: {s[1]}')

    print(f'\nDuplicates: {len(duplicates)}')
    for d in duplicates:
        print(f'  = {d[0]} | {d[1]}')

    print(f'\nTotal PDFs : {len(pdfs)}')
    print(f'Success    : {len(saved)}/{len(pdfs)}')


# ══════════════════════════════════════════════════
# FLOW 3 — SEARCH INVOICE
# Search by vendor name, invoice number, or date
# ══════════════════════════════════════════════════

def flow_search():
    if not os.path.exists(EXCEL_FILE):
        print('No Excel file found.'); return

    print('\nSearch by:')
    print('  1. Vendor name')
    print('  2. Invoice number')
    print('  3. Date')
    print('  4. Amount')
    opt = input('Choose (1/2/3/4): ').strip()

    keyword = input('Enter search keyword: ').strip().lower()
    if not keyword: return

    wb  = openpyxl.load_workbook(EXCEL_FILE)
    ws  = wb.active

    # Column index to search based on option
    col_map = {'1': 12, '2': 14, '3': 6, '4': 16}
    col     = col_map.get(opt, 8)

    results = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cell_val = str(row[col] or '').lower()
        if keyword in cell_val:
            results.append((i, row))

    if not results:
        print(f'No invoices found matching "{keyword}"')
        return

    print(f'\nFound {len(results)} result(s):\n')
    print(f'  {"Row":<5} {"Vendor":<25} {"Invoice No":<15} {"Date":<15} {"Amount":<12} {"TDS":<15} {"Status"}')
    print('  ' + '-'*95)
    for row_no, row in results:
        print(f'  {row_no:<5} {str(row[12] or ""):<25} {str(row[14] or ""):<15} '
              f'{str(row[6] or ""):<15} {str(row[16] or ""):<12} {str(row[13] or ""):<15} {str(row[19] or "")}')


# ══════════════════════════════════════════════════
# FLOW 4 — UPDATE INVOICE
# Find a row by invoice number and vendor, then correct any field
# ══════════════════════════════════════════════════

def flow_update():
    if not os.path.exists(EXCEL_FILE):
        print('No Excel file found.'); return

    vendor  = input('\nVendor name   : ').strip().lower()
    inv_no  = input('Invoice number: ').strip().lower()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    target_row = None
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if (str(row[8] or '').lower() == vendor and
                str(row[10] or '').lower() == inv_no):
            target_row = i
            break

    if not target_row:
        print(f'Invoice not found: {inv_no} from {vendor}')
        return

    print(f'\nFound at row {target_row}. Current values:')
    row_data = list(ws.iter_rows(min_row=target_row, max_row=target_row, values_only=True))[0]
    for i, (col, val) in enumerate(zip(COLS, row_data)):
        print(f'  {i+1:>2}. {col:<20}: {val or "(empty)"}')

    print('\nWhich field to update? Enter field number (or 0 to cancel):')
    field_no = input('Field number: ').strip()
    if field_no == '0' or not field_no.isdigit(): return

    field_idx = int(field_no)
    if field_idx < 1 or field_idx > len(COLS):
        print('Invalid field number.'); return

    current = row_data[field_idx - 1]
    new_val = input(f'New value for {COLS[field_idx-1]} [{current}]: ').strip()
    if not new_val:
        print('No change made.'); return

    ws.cell(target_row, field_idx, new_val)
    wb.save(EXCEL_FILE)
    print(f'Updated {COLS[field_idx-1]} to: {new_val}')


# ══════════════════════════════════════════════════
# FLOW 5 — DELETE INVOICE
# Find by invoice number + vendor, confirm, then delete row
# ══════════════════════════════════════════════════

def flow_delete():
    if not os.path.exists(EXCEL_FILE):
        print('No Excel file found.'); return

    vendor = input('\nVendor name   : ').strip().lower()
    inv_no = input('Invoice number: ').strip().lower()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    target_row = None
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if (str(row[8] or '').lower() == vendor and
                str(row[10] or '').lower() == inv_no):
            target_row = i
            break

    if not target_row:
        print(f'Invoice not found: {inv_no} from {vendor}')
        return

    row_data = list(ws.iter_rows(min_row=target_row, max_row=target_row, values_only=True))[0]
    print(f'\nFound at row {target_row}:')
    print(f'  Vendor     : {row_data[8]}')
    print(f'  Invoice No : {row_data[10]}')
    print(f'  Date       : {row_data[6]}')
    print(f'  Amount     : {row_data[12]}')
    print(f'  TDS        : {row_data[9]}')

    confirm = input('\nAre you sure you want to DELETE this invoice? (YES to confirm): ').strip().upper()
    if confirm != 'YES':
        print('Cancelled.'); return

    ws.delete_rows(target_row)
    wb.save(EXCEL_FILE)
    print(f'Invoice {inv_no} deleted from Excel.')


# ══════════════════════════════════════════════════
# FLOW 6 — MONTHLY SUMMARY
# ══════════════════════════════════════════════════

def flow_monthly_summary():
    if not os.path.exists(EXCEL_FILE):
        print('No Excel file found.'); return

    month_input = input('\nEnter month and year (e.g. December 2025): ').strip()
    wb  = openpyxl.load_workbook(EXCEL_FILE)
    ws  = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if r[2] and r[3] and f'{r[2]} {r[3]}' == month_input]

    if not rows:
        print(f'No invoices found for {month_input}'); return

    total_amt, total_tds, pending = 0, 0, []
    for r in rows:
        try: total_amt += float(str(r[16]).replace(',',''))
        except: pass
        tds_val = str(r[13]) if r[13] else ''
        if tds_val and tds_val != 'No TDS':
            try:
                rate = float(re.search(r'(\d+)%', tds_val).group(1)) / 100
                total_tds += float(str(r[25]).replace(',','')) * rate
            except: pass
        if not r[19] or str(r[19]).strip() == '':
            pending.append(f'{r[12]} - INV {r[14]} - Rs {r[16]}')

    print(f'\nMONTHLY SUMMARY: {month_input}')
    print(f'  Total Invoices   : {len(rows)}')
    print(f'  Total Amount     : Rs {round(total_amt, 2)}')
    print(f'  Total TDS        : Rs {round(total_tds, 2)}')
    print(f'  Pending Payments : {len(pending)}')
    for p in pending: print(f'    - {p}')


# ══════════════════════════════════════════════════
# FLOW 7 — VENDOR HISTORY
# ══════════════════════════════════════════════════

def flow_vendor_history():
    if not os.path.exists(EXCEL_FILE):
        print('No Excel file found.'); return

    vendor_name = input('\nEnter vendor name: ').strip()
    wb   = openpyxl.load_workbook(EXCEL_FILE)
    ws   = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
            if r[8] and vendor_name.lower() in str(r[8]).lower()]

    if not rows:
        print(f'No invoices found for {vendor_name}'); return

    print(f'\nVENDOR HISTORY: {vendor_name}')
    print(f'  {"Invoice No":<15} {"Date":<15} {"Amount":<12} {"TDS":<15} {"Status"}')
    print('  ' + '-'*70)
    total = 0
    for r in rows:
        amt = r[12] or 0
        try: total += float(str(amt).replace(',',''))
        except: pass
        print(f'  {str(r[10]):<15} {str(r[6]):<15} {str(amt):<12} {str(r[9]):<15} {str(r[13] or "")}')
    print(f'\n  Total paid to {vendor_name}: Rs {total}')


# ══════════════════════════════════════════════════
# MAIN MENU
# ══════════════════════════════════════════════════

if __name__ == '__main__':
    print('=' * 50)
    print('  INVOICE TRACKER — MakerInMe Technologies')
    print('=' * 50)
    print('  1. Add single invoice')
    print('  2. Batch process (multiple invoices)')
    print('  3. Search invoice')
    print('  4. Update invoice')
    print('  5. Delete invoice')
    print('  6. Monthly summary report')
    print('  7. Vendor history')
    print('=' * 50)

    choice = input('\nChoose (1-7): ').strip()

    if   choice == '1': flow_add_invoice()
    elif choice == '2': flow_batch()
    elif choice == '3': flow_search()
    elif choice == '4': flow_update()
    elif choice == '5': flow_delete()
    elif choice == '6': flow_monthly_summary()
    elif choice == '7': flow_vendor_history()
    else: print('Invalid choice.')
