"""
Invoice Tracker - FastAPI Backend  (Performance-Optimised v2)
─────────────────────────────────────────────────────────────
KEY FIXES vs original:
  • Tier-1 regex TDS  : instant rule-matching — no AI for ~70% invoices
  • TDS rules cache   : fetched once per 5 min, not on every call
  • Parallel warmup   : TDS rules cache warmed while PDF is being extracted
  • Reduced tokens    : 400→250 extraction, 80→60 TDS
  • Smaller images    : 420→380px width, quality 70→65
  • Scanned PDFs      : only page 1 processed (was 2)
  • LM timeout        : 600→120s — fail-fast instead of hanging for minutes
  • elapsed_sec       : returned so UI can show extraction time

Start:  python api.py
API  :  http://localhost:8000
"""

import os, re, json, base64, datetime, io, time
import asyncio, concurrent.futures
import requests as req
import openpyxl
import pdfplumber
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List
from pdf2image import convert_from_path
from PIL import Image

app = FastAPI(title="Invoice Tracker API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

MODEL        = "qwen/qwen2.5-vl-7b"
LM_URL       = "http://127.0.0.1:1234/v1/chat/completions"
TDS_API_URL  = "http://localhost:5000/tds-rules"
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE   = os.path.join(BASE_DIR, "invoice_tracker.xlsx")
IMAGES_DIR   = os.path.join(BASE_DIR, "invoice_pages")
UPLOAD_DIR   = os.path.join(BASE_DIR, "uploads")
POPPLER_PATH = r"C:\Users\chowd\Downloads\Release-25.12.0-0\poppler-25.12.0\Library\bin"

os.makedirs(IMAGES_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR,  exist_ok=True)

# ── TDS rules cache (5-minute TTL) ─────────────────────────
_tds_cache = {"rules": "", "at": 0.0}
TDS_TTL    = 300

COLS = [
    "SI No","Week","Month Name","Year","ID","Invoice / PI",
    "Invoice Date","Due Date","Project Name","Owner","Category",
    "Vendor ID","Vendor Name","TDS","Invoice Number","INV_SaveCopy",
    "Inv_Amt","Amount Paid / To be Paid","Partial Paid Amount",
    "Payment Status","Paid From","Payment Date",
    "GST Inputs Received Or Not","Material Received Date",
    "Material Received Status","Base Amount","SGST","CGST","IGST",
    "TDS Payment Date","TDS Done","Comments"
]
WIDTHS = [8,10,14,8,20,15,14,14,20,15,15,12,25,15,18,55,16,20,18,16,15,14,22,20,20,14,12,12,12,16,12,20]


# ───────────────────────────────────────────────────────────
# HELPERS
# ───────────────────────────────────────────────────────────

def clean_amount(val):
    if not val: return ""
    s = str(val).strip()
    if re.match(r"^\d+(\.\d+)?%$", s): return ""
    s = s.replace("₹","").replace("Rs","").replace("INR","").replace("$","").strip()
    if "=" in s: s = s.split("=")[-1].strip()
    if "+" in s:
        try:
            nums = re.findall(r"[\d.]+", s)
            return str(round(sum(float(n) for n in nums), 2))
        except: pass
    s = s.replace(",","").strip()
    m = re.search(r"[\d.]+", s)
    return m.group() if m else ""

def fv(data, key):
    v = data.get(key,"")
    return "" if v in (None,"NOT AVAILABLE","null","N/A") else str(v).strip()

def parse_json_safe(text):
    text = re.sub(r"```json|```","",text).strip()
    try: return json.loads(text)
    except:
        m = re.search(r"\{.*\}", text, re.DOTALL)
        return json.loads(m.group()) if m else {}

def parse_date(d):
    for f in ["%d %b %Y","%d/%m/%Y","%d-%m-%Y","%Y-%m-%d","%d %B %Y"]:
        try: return datetime.datetime.strptime(d.strip(), f)
        except: pass
    return None

def month_year(d):
    dt = parse_date(d)
    return (dt.strftime("%B"), str(dt.year)) if dt else ("","")


# ───────────────────────────────────────────────────────────
# LM STUDIO CALL
# ───────────────────────────────────────────────────────────

def call_ai(messages, image=None, tokens=250):
    if image:
        messages[-1]["content"] = [
            {"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{image}"}},
            {"type":"text","text":messages[-1]["content"]}
        ]
    r = req.post(LM_URL, json={
        "model":MODEL,"messages":messages,
        "temperature":0.0,"max_tokens":tokens,"stream":False,
        "options":{"num_predict":tokens,"top_k":1}
    }, timeout=120)
    return r.json()["choices"][0]["message"]["content"].strip()


# ───────────────────────────────────────────────────────────
# TDS RULES (cached)
# ───────────────────────────────────────────────────────────

def get_tds_rules():
    now = time.time()
    if _tds_cache["rules"] and (now - _tds_cache["at"]) < TDS_TTL:
        return _tds_cache["rules"]
    try:
        rules = req.get(TDS_API_URL, timeout=5).json().get("rules","")
    except:
        f = os.path.join(BASE_DIR,"tds_rules.txt")
        rules = open(f,encoding="utf-8").read() if os.path.exists(f) else ""
    _tds_cache["rules"] = rules
    _tds_cache["at"]    = now
    return rules


# ───────────────────────────────────────────────────────────
# REGEX TDS (instant — no AI)
# ───────────────────────────────────────────────────────────

def regex_tds(extracted):
    mat  = (fv(extracted,"material_indicator") or "").lower()
    hsn  = fv(extracted,"hsn_code")
    desc = (fv(extracted,"service_description") or "").lower()
    try:   amt_f = float((fv(extracted,"base_amount") or "0").replace(",",""))
    except: amt_f = 0.0

    mat_kw = ["vinyl","paper","ink","flex","fabric","board","sheet","plastic","metal","wood","cement","supplied","provided"]
    if any(k in mat for k in mat_kw) and "no physical" not in mat:
        return "No TDS","Vendor supplies material — product purchase, TDS not applicable","high"

    if hsn:
        if re.match(r"^(39|48|49)", hsn):
            return "No TDS",f"HSN {hsn} is a goods code — no TDS applicable","high"
        if re.match(r"^9988", hsn):
            return "194C @ 2%",f"HSN {hsn} — manufacturing on inputs owned by others","high"

    if 0 < amt_f < 30000:
        return "No TDS",f"Amount ₹{amt_f:,.0f} below ₹30,000 threshold — TDS not applicable","high"

    prof_kw = ["consulting","legal","software","accounting","design","training","it service","advisory","audit"]
    cont_kw = ["transport","courier","security","manpower","housekeeping","labour","printing","fabrication","installation"]

    if any(k in desc for k in prof_kw):
        rate = "10%" if "technical" not in desc else "2%"
        return f"194J @ {rate}","Professional/technical service detected — Sec 194J","medium"
    if any(k in desc for k in cont_kw):
        return "194C @ 2%","Contractor payment detected — Sec 194C","medium"

    return None  # inconclusive


# ───────────────────────────────────────────────────────────
# PDF EXTRACTION
# ───────────────────────────────────────────────────────────

def is_digital_pdf(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text and len(text.strip()) > 100:
                    return True
    except: pass
    return False

def extract_text_from_pdf(pdf_path):
    full = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages[:3]:
            t = page.extract_text()
            if t: full += t + "\n"
    return full.strip()

def parse_invoice_text(text):
    result = {}
    lines  = text.split("\n")
    lower  = text.lower()

    if "tax invoice" in lower:       result["document_type"] = "Tax Invoice"
    elif "proforma"  in lower:       result["document_type"] = "Proforma Invoice"
    elif "purchase order" in lower:  result["document_type"] = "Purchase Order"
    else:                            result["document_type"] = "Invoice"

    gstins = re.findall(r'\b\d{2}[A-Z]{5}\d{4}[A-Z]{1}[A-Z\d]{1}Z[A-Z\d]{1}\b', text)
    if len(gstins) >= 1: result["vendor_gstin"] = gstins[0]
    if len(gstins) >= 2: result["buyer_gstin"]  = gstins[1]

    for p in [r'invoice\s*(?:no|number|#)[:\s#]*([A-Z0-9/_-]+)',
              r'inv\s*(?:no|#)?[:\s#]*([A-Z0-9/_-]+)',
              r'bill\s*(?:no|number)[:\s#]*([A-Z0-9/_-]+)']:
        m = re.search(p, text, re.IGNORECASE)
        if m: result["invoice_number"] = m.group(1).strip(); break

    dp    = r'\b(\d{1,2}[\s/-](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s/-]\d{2,4}|\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2})\b'
    dates = re.findall(dp, text, re.IGNORECASE)
    if dates:           result["invoice_date"] = dates[0]
    if len(dates) >= 2: result["due_date"]     = dates[1]

    for line in lines[:8]:
        line = line.strip()
        if (len(line) > 4 and
            not re.search(r'invoice|tax|gstin|date|phone|email|www|http', line, re.IGNORECASE) and
            not re.search(r'^\d', line) and len(line) < 60):
            result["vendor_name"] = line; break

    for i, line in enumerate(lines):
        if re.search(r'bill\s*to|billed\s*to|ship\s*to', line, re.IGNORECASE):
            for l in lines[i+1:i+4]:
                l = l.strip()
                if len(l) > 3 and not re.search(r'gstin|gst|address', l, re.IGNORECASE):
                    result["buyer_name"] = l; break
            break

    hsn = re.search(r'hsn[/\s]*(?:sac)?[:\s]*(\d{4,8})', text, re.IGNORECASE)
    if hsn: result["hsn_code"] = hsn.group(1)
    else:
        h2 = re.findall(r'\b(\d{6,8})\b', text)
        if h2: result["hsn_code"] = h2[0]

    def find_amount(patterns):
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                val = m.group(1).replace(",","").strip()
                try: float(val); return val
                except: pass
        return ""

    result["total_with_gst"] = find_amount([
        r'(?:grand\s*total|total\s*amount|amount\s*payable|net\s*payable|total\s*due)[:\s₹Rs.]*([0-9,]+(?:\.\d{1,2})?)',
        r'(?:total)[:\s₹Rs.]*([0-9,]+(?:\.\d{1,2})?)\s*$',
    ])
    result["base_amount"] = find_amount([
        r'(?:taxable\s*(?:value|amount)|sub\s*total|base\s*amount|amount\s*before\s*tax)[:\s₹Rs.]*([0-9,]+(?:\.\d{1,2})?)',
        r'(?:total\s*taxable)[:\s₹Rs.]*([0-9,]+(?:\.\d{1,2})?)',
    ])
    result["cgst"] = find_amount([r'cgst[:\s@%\d.]*?(?:amount)?[:\s₹Rs.]*([0-9,]+\.\d{2})',r'cgst\s+[\d.]+%?\s+([0-9,]+\.\d{2})'])
    result["sgst"] = find_amount([r'sgst[:\s@%\d.]*?(?:amount)?[:\s₹Rs.]*([0-9,]+\.\d{2})',r'sgst\s+[\d.]+%?\s+([0-9,]+\.\d{2})'])
    result["igst"] = find_amount([r'igst[:\s@%\d.]*?(?:amount)?[:\s₹Rs.]*([0-9,]+\.\d{2})',r'igst\s+[\d.]+%?\s+([0-9,]+\.\d{2})'])

    for i, line in enumerate(lines):
        if re.search(r'\b(?:description|particulars|item|product|service)\b', line, re.IGNORECASE):
            for l in lines[i+1:i+4]:
                l = l.strip()
                if len(l) > 5 and not re.search(r'qty|rate|amount|total|hsn', l, re.IGNORECASE):
                    result["service_description"] = l; break
            break

    mat_kw = ["vinyl","paper","ink","flex","fabric","material","board","sheet","plastic","metal","wood","cement"]
    found  = [k for k in mat_kw if k in lower]
    result["material_indicator"] = ", ".join(found)+" material included" if found else "no physical material mentioned"

    return {k: v for k, v in result.items() if v}


def pdf_page_to_jpeg_b64(page):
    w, h = page.size
    if w > 380:
        page = page.resize((380, int(h*380/w)), Image.LANCZOS)
    buf = io.BytesIO()
    page.convert("RGB").save(buf, format="JPEG", quality=65, optimize=True)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def extract_from_pdf(pdf_path):
    result = {}
    if is_digital_pdf(pdf_path):
        pdf_text = extract_text_from_pdf(pdf_path)
        print("Digital PDF — regex extraction...")
        result = parse_invoice_text(pdf_text)
        missing = [f for f in ["vendor_name","invoice_number","invoice_date","total_with_gst"] if not result.get(f)]
        if missing:
            print(f"Regex missed {missing} — AI text mode...")
            data = parse_json_safe(call_ai(
                [{"role":"user","content":
                    'Extract invoice data. Return ONLY JSON:\n'
                    '{"document_type":"","vendor_name":"","vendor_gstin":"","invoice_number":"",'
                    '"invoice_date":"","due_date":"","buyer_name":"","buyer_gstin":"","hsn_code":"",'
                    '"service_description":"","material_indicator":"","base_amount":"","cgst":"",'
                    '"sgst":"","igst":"","total_with_gst":""}\n'
                    'Dates:dd mmm yyyy. Amounts:numbers only. igst:empty if absent.\n\n'
                    f'TEXT:\n{pdf_text[:1800]}'}],
                tokens=250
            ))
            for k, v in data.items():
                if v and v not in ("","NOT AVAILABLE","null","N/A"):
                    result.setdefault(k, v)
        else:
            print("Regex OK — no AI needed!")
    else:
        print("Scanned PDF — AI vision (page 1)...")
        pages   = convert_from_path(pdf_path, dpi=80, poppler_path=POPPLER_PATH)
        img_b64 = pdf_page_to_jpeg_b64(pages[0])
        data    = parse_json_safe(call_ai(
            [{"role":"user","content":
                'Extract from this invoice image. Return ONLY JSON:\n'
                '{"document_type":"","vendor_name":"","vendor_gstin":"","invoice_number":"",'
                '"invoice_date":"","due_date":"","buyer_name":"","buyer_gstin":"","hsn_code":"",'
                '"service_description":"","material_indicator":"","base_amount":"","cgst":"",'
                '"sgst":"","igst":"","total_with_gst":""}\n'
                'vendor=TOP. buyer=Bill To. Dates:dd mmm yyyy. Amounts:numbers only. igst:empty if absent.'}],
            image=img_b64, tokens=250
        ))
        result = {k:v for k,v in data.items() if v and v not in ("","NOT AVAILABLE","null","N/A")}

    for f in ["base_amount","cgst","sgst","igst","total_with_gst"]:
        if f in result: result[f] = clean_amount(result[f])
    return result


# ───────────────────────────────────────────────────────────
# TDS DECISION  (regex-first, AI fallback)
# ───────────────────────────────────────────────────────────

def auto_tds(extracted):
    r = regex_tds(extracted)
    if r:
        print(f"TDS via regex: {r[0]}")
        return r
    print("TDS inconclusive — asking AI...")
    result = parse_json_safe(call_ai([
        {"role":"system","content":get_tds_rules()},
        {"role":"user","content":
            f"Vendor:{fv(extracted,'vendor_name')} "
            f"Service:{fv(extracted,'service_description')} "
            f"HSN:{fv(extracted,'hsn_code')} "
            f"Materials:{fv(extracted,'material_indicator')} "
            f"Amount:Rs {fv(extracted,'base_amount')}\n"
            'Reply ONLY as JSON: {"tds_value":"","reason":"","confidence":"high/medium/low"}'}
    ], tokens=60))
    return result.get("tds_value",""), result.get("reason",""), result.get("confidence","low")


# ───────────────────────────────────────────────────────────
# EXCEL UTILITIES
# ───────────────────────────────────────────────────────────

def confidence_score(data):
    keys   = ["vendor_name","invoice_number","invoice_date","base_amount","total_with_gst","hsn_code"]
    filled = sum(1 for k in keys if fv(data,k))
    score  = filled / len(keys)
    if score >= 0.8:   return "High",   []
    elif score >= 0.5: return "Medium", [k for k in keys if not fv(data,k)]
    else:              return "Low",    [k for k in keys if not fv(data,k)]

def check_duplicate(vendor, invoice_no):
    if not os.path.exists(EXCEL_FILE): return False
    ws = openpyxl.load_workbook(EXCEL_FILE).active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[12] or "").lower()==vendor.lower() and str(row[14] or "").lower()==invoice_no.lower():
            return True
    return False

def get_or_create_wb():
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE), openpyxl.load_workbook(EXCEL_FILE).active
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Invoice Tracker"
    for i, c in enumerate(COLS, 1):
        cell = ws.cell(1, i, c)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.row_dimensions[1].height = 30
    wb.save(EXCEL_FILE)
    return wb, ws

def get_or_create_wb():
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        return wb, wb.active
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Invoice Tracker"
    for i,c in enumerate(COLS,1):
        cell = ws.cell(1,i,c)
        cell.font = Font(name="Calibri",bold=True,size=10)
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    for i,w in enumerate(WIDTHS,1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.row_dimensions[1].height = 30
    wb.save(EXCEL_FILE); return wb, ws

def save_row(data):
    wb, ws = get_or_create_wb()
    row_no = ws.max_row + 1; si = ws.max_row
    dt = data.get("invoice_date",""); mn,yr = month_year(dt)
    uid = data.get("unique_id",""); v = data.get("vendor_name","")
    inv = data.get("invoice_number",""); tot = data.get("total_amount","")
    copy = f"{uid} {dt} {v} {inv} {tot}"
    values = [si,data.get("week",""),mn,yr,uid,data.get("doc_type",""),dt,
        data.get("due_date",""),data.get("project_name",""),data.get("owner",""),
        data.get("category",""),data.get("vendor_id",""),v,data.get("tds",""),
        inv,copy,tot,data.get("amount_paid",""),data.get("partial_paid",""),
        data.get("payment_status",""),data.get("paid_from",""),data.get("payment_date",""),
        data.get("gst_inputs",""),data.get("material_received_date",""),
        data.get("material_received_status",""),data.get("base_amount",""),
        data.get("sgst",""),data.get("cgst",""),data.get("igst",""),
        data.get("tds_payment_date",""),data.get("tds_done",""),data.get("comments","")]
    for i,val in enumerate(values,1):
        cell = ws.cell(row_no,i,val)
        cell.font = Font(name="Calibri",size=10)
        cell.alignment = Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[row_no].height = 18
    wb.save(EXCEL_FILE); return si, copy


# ───────────────────────────────────────────────────────────
# ENDPOINTS
# ───────────────────────────────────────────────────────────

@app.get("/health")
def health():
    lm_ok = tds_ok = False
    try:
        r = req.post(LM_URL,json={"model":MODEL,"messages":[{"role":"user","content":"hi"}],"max_tokens":5,"stream":False},timeout=10)
        lm_ok = "choices" in r.json()
    except: pass
    try:
        r = req.get(TDS_API_URL,timeout=5); tds_ok = r.json().get("status")=="success"
    except: pass
    return {"lm_studio":lm_ok,"tds_api":tds_ok}


@app.post("/invoice/extract")
async def extract_invoice(file: UploadFile = File(...)):
    pdf_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(pdf_path,"wb") as f: f.write(await file.read())
    try:
        t0      = time.time()
        digital = is_digital_pdf(pdf_path)

        # Warm TDS cache in background while extraction runs
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
            tds_warm = loop.run_in_executor(pool, get_tds_rules)
            ext_fut  = loop.run_in_executor(pool, extract_from_pdf, pdf_path)
            await tds_warm        # ensures cache is ready before auto_tds
            extracted = await ext_fut

        if not extracted:
            raise HTTPException(status_code=422, detail="Could not read invoice")

        tds_val, tds_why, tds_conf = auto_tds(extracted)
        conf, missing = confidence_score(extracted)

        due_alert = ""
        due_str = fv(extracted,"due_date")
        if due_str:
            dt = parse_date(due_str)
            if dt:
                diff = (datetime.datetime.today()-dt).days
                if diff > 0:     due_alert = f"OVERDUE by {diff} day(s)"
                elif diff == 0:  due_alert = "Due TODAY"
                elif diff >= -3: due_alert = f"Due in {abs(diff)} day(s)"

        elapsed = round(time.time()-t0, 1)
        print(f"✓ Done in {elapsed}s")
        return {
            "extracted":extracted,"confidence":conf,"missing":missing,
            "tds_value":tds_val,"tds_reason":tds_why,"tds_conf":tds_conf,
            "due_alert":due_alert,
            "is_duplicate":check_duplicate(fv(extracted,"vendor_name"),fv(extracted,"invoice_number")),
            "pdf_type":"digital" if digital else "scanned",
            "elapsed_sec":elapsed,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class InvoiceData(BaseModel):
    unique_id:str=""; doc_type:str=""; vendor_name:str=""; vendor_gstin:str=""
    invoice_number:str=""; invoice_date:str=""; due_date:str=""; buyer_name:str=""
    buyer_gstin:str=""; hsn_code:str=""; service_desc:str=""; base_amount:str=""
    cgst:str=""; sgst:str=""; igst:str=""; total_amount:str=""; tds:str=""
    week:str=""; project_name:str=""; owner:str=""; category:str=""; vendor_id:str=""
    amount_paid:str=""; partial_paid:str=""; payment_status:str=""; paid_from:str=""
    payment_date:str=""; gst_inputs:str=""; material_received_date:str=""
    material_received_status:str=""; tds_payment_date:str=""; tds_done:str=""; comments:str=""

@app.post("/invoice/save")
def save_invoice(data: InvoiceData):
    try:
        si, copy = save_row(data.dict())
        return {"success":True,"row":si,"invoice_copy":copy}
    except PermissionError:
        raise HTTPException(status_code=423,detail="Excel file is open. Please close it.")
    except Exception as e:
        raise HTTPException(status_code=500,detail=str(e))

@app.post("/invoice/batch")
async def batch_process(files: List[UploadFile] = File(...)):
    saved,skipped,duplicates=[],[],[]
    for file in files:
        pdf_path = os.path.join(UPLOAD_DIR,file.filename)
        with open(pdf_path,"wb") as f: f.write(await file.read())
        try:
            extracted = extract_from_pdf(pdf_path)
            if not extracted: skipped.append({"file":file.filename,"reason":"Could not read"}); continue
            conf,missing = confidence_score(extracted)
            if conf=="Low": skipped.append({"file":file.filename,"reason":f"Low confidence - missing: {', '.join(missing)}"}); continue
            vendor=fv(extracted,"vendor_name"); inv_no=fv(extracted,"invoice_number")
            if check_duplicate(vendor,inv_no): duplicates.append({"file":file.filename,"vendor":vendor,"invoice":inv_no}); continue
            tds_val,tds_why,tds_conf = auto_tds(extracted)
            if tds_conf=="low": tds_val=f"[NEEDS REVIEW] {tds_val}"
            uid=f"{inv_no}_{fv(extracted,'invoice_date').replace(' ','')}" or "AUTO"
            data={"unique_id":uid,"doc_type":fv(extracted,"document_type"),"vendor_name":vendor,
                "vendor_gstin":fv(extracted,"vendor_gstin"),"invoice_number":inv_no,
                "invoice_date":fv(extracted,"invoice_date"),"due_date":fv(extracted,"due_date"),
                "buyer_name":fv(extracted,"buyer_name"),"hsn_code":fv(extracted,"hsn_code"),
                "service_desc":fv(extracted,"service_description"),
                "base_amount":clean_amount(fv(extracted,"base_amount")),
                "cgst":clean_amount(fv(extracted,"cgst")),"sgst":clean_amount(fv(extracted,"sgst")),
                "igst":clean_amount(fv(extracted,"igst")),
                "total_amount":clean_amount(fv(extracted,"total_with_gst")),
                "tds":tds_val,"comments":f"Auto-processed. Confidence:{conf}. {tds_why}"}
            save_row(data)
            saved.append({"file":file.filename,"vendor":vendor,"invoice":inv_no,"tds":tds_val})
        except Exception as e:
            skipped.append({"file":file.filename,"reason":str(e)})
    return {"total":len(files),"saved":saved,"skipped":skipped,"duplicates":duplicates}

@app.get("/invoice/all")
def get_all():
    if not os.path.exists(EXCEL_FILE): return {"invoices":[]}
    ws=openpyxl.load_workbook(EXCEL_FILE).active
    invoices=[]
    for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        if any(v for v in row):
            invoices.append({"row":i,"si":row[0],"vendor":row[12],"invoice_number":row[14],
                "date":row[6],"due_date":row[7],"amount":row[16],"tds":row[13],"status":row[19],
                "base_amount":row[25],"cgst":row[27],"sgst":row[26]})
    return {"invoices":invoices}

@app.get("/invoice/search")
def search(q:str, field:str="vendor"):
    if not os.path.exists(EXCEL_FILE): return {"results":[]}
    col_map={"vendor":12,"invoice":14,"date":6,"amount":16}
    col=col_map.get(field,12)
    ws=openpyxl.load_workbook(EXCEL_FILE).active
    results=[]
    for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
        if str(row[col] or "").lower().find(q.lower())!=-1:
            results.append({"row":i,"vendor":row[12],"invoice_number":row[14],
                "date":row[6],"amount":row[16],"tds":row[13],"status":row[19]})
    return {"results":results}

@app.get("/invoice/summary/{month}/{year}")
def summary(month:str,year:str):
    if not os.path.exists(EXCEL_FILE): return {"total_invoices":0,"total_amount":0,"total_tds":0,"pending":[]}
    ws=openpyxl.load_workbook(EXCEL_FILE).active
    rows=[r for r in ws.iter_rows(min_row=2,values_only=True)
          if r[2] and r[3] and str(r[2]).lower()==month.lower() and str(r[3])==year]
    total_amt=total_tds=0; pending=[]
    for r in rows:
        try: total_amt+=float(str(r[16]).replace(",",""))
        except: pass
        tds_v=str(r[13] or "")
        if tds_v and tds_v!="No TDS":
            try:
                rate=float(re.search(r"(\d+)%",tds_v).group(1))/100
                total_tds+=float(str(r[25]).replace(",",""))*rate
            except: pass
        if not r[19] or str(r[19]).strip()=="":
            pending.append({"vendor":r[12],"invoice":r[14],"amount":r[16]})
    return {"total_invoices":len(rows),"total_amount":round(total_amt,2),
            "total_tds":round(total_tds,2),"pending":pending,"pending_count":len(pending)}

@app.get("/invoice/vendor/{vendor_name}")
def vendor_hist(vendor_name:str):
    if not os.path.exists(EXCEL_FILE): return {"invoices":[],"total":0}
    ws=openpyxl.load_workbook(EXCEL_FILE).active
    rows=[r for r in ws.iter_rows(min_row=2,values_only=True)
          if r[12] and vendor_name.lower() in str(r[12]).lower()]
    total=0; invoices=[]
    for r in rows:
        try: total+=float(str(r[16] or 0).replace(",",""))
        except: pass
        invoices.append({"invoice_number":r[14],"date":r[6],"amount":r[16],"tds":r[13],"status":r[19]})
    return {"invoices":invoices,"total":round(total,2),"vendor":vendor_name}

class UpdateData(BaseModel):
    row:int; field:str; value:str

@app.put("/invoice/update")
def update(data:UpdateData):
    col_map={"vendor_name":13,"invoice_number":15,"invoice_date":7,"due_date":8,"amount":17,
             "tds":14,"payment_status":20,"paid_from":21,"payment_date":22,"comments":32,
             "tds_done":31,"base_amount":26,"cgst":28,"sgst":27,"igst":29}
    col=col_map.get(data.field)
    if not col: raise HTTPException(status_code=400,detail=f"Unknown field: {data.field}")
    wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb.active
    ws.cell(data.row,col,data.value); wb.save(EXCEL_FILE)
    return {"success":True}

@app.delete("/invoice/delete/{row}")
def delete(row:int):
    wb=openpyxl.load_workbook(EXCEL_FILE); ws=wb.active
    ws.delete_rows(row); wb.save(EXCEL_FILE)
    return {"success":True}

@app.get("/invoice/download")
def download():
    if not os.path.exists(EXCEL_FILE): raise HTTPException(status_code=404,detail="No Excel file")
    return FileResponse(EXCEL_FILE,filename="invoice_tracker.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__":
    import uvicorn
    uvicorn.run("api:app",host="0.0.0.0",port=8000,reload=True)
